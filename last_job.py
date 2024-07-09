import requests
import openpyxl
from datetime import datetime

# Function to get the last job executed date from Jenkins
def get_last_job_executed_date(url, username, api_token):
    try:
        # Jenkins API endpoint to list all jobs
        api_url = f"{url}/api/json?tree=jobs[name,lastBuild[timestamp]]"
        response = requests.get(api_url, auth=(username, api_token))
        response.raise_for_status()
        
        jobs = response.json().get('jobs', [])
        if not jobs:
            return 'No Jobs Found'
        
        # Find the latest job execution date
        last_execution_time = None
        for job in jobs:
            last_build = job.get('lastBuild')
            if last_build:
                timestamp = last_build.get('timestamp')
                if timestamp:
                    execution_time = datetime.fromtimestamp(timestamp / 1000.0)
                    if not last_execution_time or execution_time > last_execution_time:
                        last_execution_time = execution_time
        
        if last_execution_time:
            return last_execution_time.strftime('%Y-%m-%d %H:%M:%S')
        else:
            return 'No Executions Found'
    except requests.RequestException as e:
        return 'Error'

# Path to your Excel file
excel_file_path = 'path/to/your/excel/file.xlsx'

# Jenkins credentials
username = "your-username"
api_token = "your-api-token"

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Add a new column header for the last job executed date if it doesn't exist
last_job_column = sheet.max_column + 1
sheet.cell(row=1, column=last_job_column, value='Last Job Executed Date')

# Iterate over each row in the sheet (starting from the second row)
for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1), start=2):
    jenkins_url = row[0].value
    if jenkins_url:
        last_job_date = get_last_job_executed_date(jenkins_url, username, api_token)
        # Save the last job executed date in the new column
        sheet.cell(row=row_index, column=last_job_column, value=last_job_date)

# Save the workbook
wb.save(excel_file_path)

print("Last job executed dates have been checked and updated in the Excel file.")
