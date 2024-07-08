import openpyxl
import subprocess
import json

# Path to your Excel file
excel_file_path = 'path\\to\\your\\file.xlsx'

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Add a new column header for SPOC if it doesn't exist
if sheet.cell(row=1, column=6).value is None:
    sheet.cell(row=1, column=6, value='SPOC')

# Function to execute the curl command and extract manager_email
def execute_curl_command(trigram, iappliid):
    curl_command = f'curl -X POST "http://your.api.endpoint" -d "trigram={trigram}&iappliid={iappliid}"'
    try:
        result = subprocess.run(curl_command, shell=True, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"Failed to execute curl for trigram={trigram} and iappliid={iappliid}")
            print(f"Error: {result.stderr}")
            return None
        else:
            response_json = json.loads(result.stdout)
            manager_email = response_json.get('manager_email')
            print(f"Successfully executed curl for trigram={trigram} and iappliid={iappliid}")
            return manager_email
    except Exception as e:
        print(f"Exception occurred: {str(e)}")
        return None

# Iterate over each row in the sheet (starting from the second row)
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):  # Assuming the first row is the header
    trigram = row[0].value
    iappliid = row[1].value
    
    # Execute the curl command and get manager_email
    manager_email = execute_curl_command(trigram, iappliid)
    
    # Save manager_email in the SPOC column (6th column)
    if manager_email:
        sheet.cell(row=row_index, column=6, value=manager_email)

# Save the workbook
wb.save(excel_file_path)

print("All API calls have been made and SPOC column has been updated.")
