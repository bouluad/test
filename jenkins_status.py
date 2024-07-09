import subprocess
import openpyxl

# Function to check if a Jenkins server is up using curl
def check_jenkins_status(url):
    try:
        # Execute the curl command
        result = subprocess.run(
            ["curl", "-o", "/dev/null", "-s", "-w", "%{http_code}", url],
            capture_output=True,
            text=True
        )
        # Get the HTTP status code from the curl output
        status_code = int(result.stdout.strip())
        if status_code == 503 or status_code == 500:
            return 'Down'
        elif status_code == 200:
            return 'Up'
        else:
            return 'Something Else'
    except Exception as e:
        return 'Error'

# Path to your Excel file
excel_file_path = 'path/to/your/excel/file.xlsx'

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Determine the last column (for status)
last_column = sheet.max_column + 1
sheet.cell(row=1, column=last_column, value='Status')

# Iterate over each row in the sheet (starting from the second row)
for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1), start=2):
    jenkins_url = row[0].value
    if jenkins_url:
        status = check_jenkins_status(jenkins_url)
        # Save the status in the last column
        sheet.cell(row=row_index, column=last_column, value=status)

# Save the workbook
wb.save(excel_file_path)

print("Jenkins server statuses have been checked and updated in the Excel file.")
